# from app.v1.models import datamodel
from re import S
import time
import os
from dateutil.parser import parse

import json
from collections import OrderedDict
from openpyxl import load_workbook
from itertools import islice
from openpyxl_image_loader import SheetImageLoader
import pandas as pd


import cloudinary as Cloud
from sqlalchemy.sql.expression import except_
# from sqlalchemy import inspect
from werkzeug.utils import secure_filename

from cloudinary.uploader import upload
from cloudinary.utils import cloudinary_url

from flask_mail import Message
from flask_login import login_required, current_user
from flask_restful import Resource, abort
from flask import render_template,Response,request,flash,redirect,url_for,make_response,jsonify
from dateutil.relativedelta import relativedelta
from sqlalchemy import exc
from app.v1.models.operations import *
from .helperfuncs import *
from .advanta import *
from operator import add
from app import sms
from app import mail

from flask_cors import cross_origin
# from rq import Queue
# from rq.job import Job
# from worker import conn
# q = Queue(connection=conn)

Cloud.config.update = ({
    'cloud_name':os.environ.get('CLOUDINARY_CLOUD_NAME'),
    'api_key': os.environ.get('CLOUDINARY_API_KEY'),
    'api_secret': os.environ.get('CLOUDINARY_API_SECRET')
})



class BComStats(Resource):
    @cross_origin()
    @login_required
    def get(self):
        data = request.get_json()
        print(data)
        if data:
            com= data.get('com')
        else:
            com="All clients"
        
        com = request.args.get('com')

        time = datetime.datetime.now()
       

        present_month = time.month
        present_year = time.year

            
        monthly_collection=[] #list of amounts from all payment objs
        client_balances=[]
        monthly_bill_members=[] #list of monthlybill amounts from monthly charge objs

        clientlist = CompanyOp.fetch_all_companies()

        print("ALL CLIENTS LISTED:")
        for cl in clientlist:
            print("COMPANY"," ",cl.id," ",cl.name," ","ACTIVE?",cl.active)

        com_obj = CompanyOp.fetch_company_by_name(com)

        clients = []

        if com == "All clients":
            comonfocus = "All Clients"
            clients = clientlist
        elif com_obj:
            if com_obj in clientlist:
                comonfocus = fname_extracter(str(com_obj))
                clients.append(com_obj)
            else:
                comonfocus = fname_extracter(str(clientlist[0]))
                clients.append(clientlist[0])
        else:
            comonfocus = fname_extracter(str(clientlist[0]))
            clients = clientlist

        dashboard_coms = []

        for i in clientlist:
            dashboard_coms.append(i.name)

        for client in clients:
            
            db.session.expire(client)

            bal_item = client.balance
            if bal_item:
                client_balances.append(bal_item)

            #######################################
            monthly_bills = client.bills
            for item in monthly_bills:
                if item.month == present_month and item.year == present_year:
                    sum_member = item.total
                    monthly_bill_members.append(sum_member)

            payment_collection = client.payments
            for item in payment_collection:
                if item.pay_period.month == present_month and item.pay_period.year == present_year:
                    sum_member = item.amount
                    monthly_collection.append(sum_member)             


        ############################################################
        monthly_bill_total = sum_positive_values(monthly_bill_members)
        monthly_total = sum_positive_values(monthly_collection)
        monthlybal_total = sum_positive_values(client_balances)
        ############################################################    
        formatted_monthly_bill_total = (f"{monthly_bill_total:,.1f}")    
        formatted_monthly_total = (f"{monthly_total:,}")
        formatted_monthlybal_total = (f"{monthlybal_total:,}")
        ############################################################


        data = {
            "month_string":get_str_mnth(present_month),
            "monthly_bills":formatted_monthly_bill_total,
            "monthly_rent_collection":formatted_monthly_total,
            "monthly_bal":formatted_monthlybal_total,
            "comonfocus":comonfocus,
            "clients" : dashboard_coms
        }
        usergroup = current_user.user_group
        user_group = str(usergroup)

        if user_group != "Admin":
            pass

        else:
            dashboard_coms.append("All clients")       
            
        return make_response(jsonify({
                "message": "Success",
                "data":data
            }), 200)    

class BComGraphStats(Resource):
    @cross_origin()
    @login_required
    def get(self):

        data = request.get_json()
        if data:
            com= data.get('com')
        else:
            com="All clients"

        time = datetime.datetime.now()
                   
        monthly_collection=[] #list of amounts from all payment objs

        clientlist = CompanyOp.fetch_all_companies()

        com_obj = CompanyOp.fetch_company_by_name(com)

        users = len(UserOp.fetch_all_users())

        clients = []
        active_clients = []
        props = []
        houses = []
        tenants = []
        active_tenants = []
        sms_spent = 0
        actual_cost = 0.0


        if com == "All clients":
            clients = clientlist
        elif com_obj:
            if com_obj in clientlist:
                clients.append(com_obj)
            else:
                clients.append(clientlist[0])
        else:

            clients.append(clientlist[0])


        for client in clients:
            # total_tenants = 0 #for setting sms quota only
            db.session.expire(client)

            props.append(client.props)

            if client.active:
                active_clients.append(client)
                       
            for prop in client.props:
                prop_tenants = tenantauto(prop.id)
                active_tenants.append(prop_tenants)
                # total_tenants += len(prop_tenants) #for setting sms quota only
                houses.append(prop.houses)
                tenants.append(prop.tenants)

            # CompanyOp.set_smsquota(client,total_tenants * 4) #for setting sms quota only

            try:
                co_smsspent = client.smsquota - client.remainingsms
                # co_smsspent += 50 #URGENT TODO REMOVE THIS 50
                sms_spent += co_smsspent
                actual_spent = co_smsspent * 0.8
                actual_cost += actual_spent
            except Exception as e:
                print("This company has", e ,"problems",client)

            payment_collection = client.payments
            for item in payment_collection:
                if item.pay_period.month == time.month and item.pay_period.year == time.year:
                    sum_member = item.amount
                    monthly_collection.append(sum_member) 
                    
        chart_string = ','.join(map(str, monthly_collection))   

        data = {
            "chartstring":chart_string,
            "clients" : len(clients),
            "active_clients" : len(active_clients),
            "registered_users": users,
            "props" : len(flatten(props)),
            "houses" : len(flatten(houses)),
            "tenants" : len(flatten(tenants)),
            "active_tenants" : len(flatten(active_tenants)),
            "smsestimate" : len(flatten(active_tenants))*3,
            "smssent" : sms_spent,
            "actualcosts" : f"Kshs {actual_cost:,.1f}"

        }

        return make_response(jsonify({
                "message": "Success",
                "data":data
            }), 200)    
            
    
            

class BRegisterOwner(Resource):
    """This class registers an owner."""
    @login_required
    def post(self):
        """ Handle POST request for this view. Url ---> /add/owner """
        data = request.get_json()
        if not data:
            return jsonify({'msg': 'Missing JSON'}), 400
        
        email= data.get('email')
        fname= data.get('fname')
        lname= data.get('lname')
        ftel= data.get('ftel')
        ltel= data.get('ltel')


        created_by = current_user.id
        name = fname+" "+lname
        tel_validation = ValidatePass.validate_password(ftel,ltel)
        if not tel_validation:
            return make_response(jsonify({
                "message": "Please provide owners' mobile number"
            }), 400)    
        if tel_validation == "no match":
            return make_response(jsonify({
                "message": "Please check the mobile number and try again"     
            }), 400) 
        else:
            phone = ftel


        uniquename = uniquename_generator(name,phone)
        is_present  = OwnerOp.fetch_owner_by_uniquename(uniquename)
        is_present2  = OwnerOp.fetch_owner_by_phone(phone)
        
        
        if is_present:
          
            return make_response(jsonify({
                "message": "Record exists in the database already"
            }), 200)      


        if is_present2:
            return make_response(jsonify({
                "message": "Owner with similar mobile number already registered"
            }), 200)      


        owner = OwnerOp(name,phone,email,uniquename,created_by)
        owner.save()

        owner_user = UserOp.fetch_user_by_phone(phone)
        
        if owner_user:
            natid = owner_user.national_id
            OwnerOp.update_natid(owner,natid)
            UserOp.update_status(owner_user,True)

        msg='Success, add apartments below.'
        flash(msg,"success")
        return make_response(jsonify({
                "message": "Success, add apartments next",
                "data":data
            }), 200)      



class BCreateApartment(Resource):
    @cross_origin()
    @login_required
    def get(self):
        owners = OwnerOp.fetch_all_owners()
        location_list = fetch_all_locations()
        regions = stringify_list_items(location_list)
        owner = stringify_list_items(owners)
        regions.sort()
    
        data={
           "owners":owner,
            "option_list":regions,
            "name":current_user.name,
            "logopath":logo(current_user.company)[0],
            "mobilelogopath":logo(current_user.company)[1]
          
        }
        return make_response(jsonify({
            "message": "Success",
            "data":data
         
        }), 200)
    
    @cross_origin()
    @login_required
    def post(self):

        user_group = current_user.company_user_group
        accessright = check_accessright(user_group,"add_apartment")
        if accessright != True:
            return make_response(jsonify({
                'msg': 'You have insufficient rights to access this form!',
                "name":current_user.name
            }))


        location_list = fetch_all_locations()
        regions = stringify_list_items(location_list)
        regions.sort()

        formatted_url = None
        name = request.form.get("name")
        owner = request.form.get("owner")
        location = request.form.get("location")
        file_to_upload = request.files.get("image") # get uploaded image

       
        if file_to_upload:
            upload_result = upload(file_to_upload) # send image to cloud
            # style the image and get its url after styling
            formatted_url, options = cloudinary_url(upload_result['public_id'],format="png",crop="fill",width=64,height=64)
            print(formatted_url)
        secure_image  = url_security(formatted_url)

        # pipeshelf = Cloud.CloudinaryImage("'upload_result[public_id]'+'.png'")
        agency_managed = request.form.get('agency_management')

        if not agency_managed:
            agency_managed = "False"

        bool_value = return_bool(agency_managed)

        # owner_id = get_owner_id(owner)
    
        return 'step'


        
        # location_id = get_location_id(location)
        
        # present = ApartmentOp.fetch_apartment_by_name(name)
        # if present:
        #     flash("Similar apartment exists","fail")
            
        #     return make_response(jsonify({
        #         "message": "Similar apartment exists",
        #     }), 200)    
            

        
        # apartment_obj = ApartmentOp(name,secure_image,location_id,owner_id,bool_value,current_user.id)
        # apartment_obj.save()
        # owner_obj = OwnerOp.fetch_owner_by_uniquename(owner)
        # owner_natid = owner_obj.national_id
        # if owner_natid:
        #     owner_user = UserOp.fetch_user_by_national_id(owner_natid)
        #     ApartmentOp.relate(apartment_obj,owner_user)

        #     if not bool_value:
        #         owner_co_id = owner_user.company_id
        #         ApartmentOp.update_company(apartment_obj,owner_co_id)

        # return make_response(jsonify({
        #         "message": "Property registration success, time to add some houses",
        #     }), 200)    
            

            


class BPropertyAccess(Resource):
    """class"""
    @cross_origin()
    @login_required
    def get(self):
        accessright = current_user.username
        if accessright != "admin":
            return make_response(jsonify({
                'msg': 'You have insufficient rights to access this form!',
                "name":current_user.name
            }))
        

        owners = OwnerOp.fetch_all_owners()
        agents = fetch_all_agents()

        

        return make_response(jsonify({
                'msg': 'Success',
                'name':current_user.name,
              'owners':stringify_list_items(owners),
            'logopath':logo(current_user.company)[0],
            'mobilelogopath':logo(current_user.company)[1],
            'agents':stringify_list_items(agents)
            }))

    @cross_origin()
    @login_required
    def post(self):
        
        data = request.get_json()
        if not data:
            return jsonify({'msg': 'Missing JSON'}), 400
    
   
        owner = data.get('owner')
        propertyauto = data.get('propertyauto')
        username = data.get('agent')

        print(OwnerOp.fetch_owner_by_uniquename(owner))
        return 'username'
    
        # owner_id = OwnerOp.fetch_owner_by_uniquename(owner).id
        # apartments = ApartmentOp.fetch_all_apartments_by_owner(owner_id)

        # apartment = data.get('property')
        
        # if propertyauto:
        #     apartments.append("All")
        #     return make_response(jsonify({
        #         'msg': 'Success',
        #       'apartmentlist':stringify_list_items(apartments)
        #     }))
        

        # agent_obj = UserOp.fetch_user_by_username(username)

        # if apartment == "All":
        #     for prop in apartments:
        #         current_apartments = fetch_all_apartments_by_user(agent_obj)
        #         if prop in current_apartments:
        #             print("skipping...")

        #         else:
        #             ApartmentOp.relate(prop,agent_obj)
        #             print("Agent given access to ",str(prop))
        #             UserOp.update_status(agent_obj,True)
        #             ApartmentOp.update_agent(prop,agent_obj.username)
        #             if prop.agency_managed:
        #                 agent_co = agent_obj.company
        #                 ApartmentOp.update_company(prop,agent_co.id)

        #                 company_users = agent_co.users
        #                 for i in company_users:
        #                     if i.user_group_id == 4:
        #                         ApartmentOp.relate(prop,i)
        #                         print("user added to ",str(prop))
        #                         return make_response(jsonify({
        #                             'msg':msg,
      
        #                         }))

        # else:
        #     apartment_obj = ApartmentOp.fetch_apartment_by_name(apartment)
        #     current_apartments = fetch_all_apartments_by_user(agent_obj)
        #     if apartment_obj in current_apartments:
        #         print("Agent already has access to ",str(apartment_obj))
        #         return make_response(jsonify({
        #             'msg':"Agent already has access"
      
        #         }))
        #     else:
        #         ApartmentOp.relate(apartment_obj,agent_obj)
        #         print("Agent given access to ",str(apartment_obj))
        #         UserOp.update_status(agent_obj,True)
        #         if not apartment_obj.agent_id:
        #             ApartmentOp.update_agent(apartment_obj,agent_obj.username)
        #             if apartment_obj.agency_managed:
        #                 agent_co = agent_obj.company
        #                 ApartmentOp.update_company(apartment_obj,agent_co.id)
    
        #                 company_users = agent_co.users
        #                 for i in company_users:
        #                     if i.user_group_id == 4:
        #                         ApartmentOp.relate(apartment_obj,i)
        #                         print("user added to apartment")


        # msg = "Operation completed"
        # flash(msg,"success")
  
        # return make_response(jsonify({
        #         'msg':msg,
      
        #     }))
        


class BPropertyAccessTermination(Resource):
    """class"""
    @login_required
    def get(self):
        accessright = current_user.username
        if accessright != "admin":
            return make_response(jsonify({
                'msg': 'You have insufficient rights to access this form!',
                "name":current_user.name
            }))
        

        owners = OwnerOp.fetch_all_owners()
        agents = fetch_all_agents()

        return make_response(jsonify({
                'msg': 'Success',
                'name':current_user.name,
              'owners':stringify_list_items(owners),
            'logopath':logo(current_user.company)[0],
            'mobilelogopath':logo(current_user.company)[1],
            'agents':stringify_list_items(agents)
            }))


    @login_required
    def post(self):

        data = request.get_json()
        if not data:
            return jsonify({'msg': 'Missing JSON'}), 400
    
   
        owner = data.get('owner')
        propertyauto = data.get('propertyauto')
        username = data.get('agent')
    
        owner_id = OwnerOp.fetch_owner_by_uniquename(owner).id
        apartments = ApartmentOp.fetch_all_apartments_by_owner(owner_id)

        apartment = data.get('property')


        owner_id = OwnerOp.fetch_owner_by_uniquename(owner).id
        apartments = ApartmentOp.fetch_all_apartments_by_owner(owner_id)

        if propertyauto:
            apartments.append("All")
            return make_response(jsonify({
                'msg': 'Success',
              'apartmentlist':stringify_list_items(apartments)
            }))
        

        agent_obj = UserOp.fetch_user_by_username(username)

        if apartment == "All":
            # for prop in apartments:
            #     current_apartments = fetch_all_apartments_by_user(agent_obj)
            #     if prop in current_apartments:
            #         print("skipping...")
            #     else:
            #         ApartmentOp.relate(prop,agent_obj)
            #         ApartmentOp.update_agent(prop,agent_obj.username)
            #         if prop.agency_managed:
            #             agent_obj_co = agent_obj.company_id
            #             ApartmentOp.update_company(prop,agent_obj_co)
            pass
        else:
            apartment_obj = ApartmentOp.fetch_apartment_by_name(apartment)
            current_apartments = fetch_all_apartments_by_user(agent_obj)
            if apartment_obj not in current_apartments:
                print("Agent did not have access to ",str(apartment_obj))
            else:
                ApartmentOp.terminate(apartment_obj,agent_obj)
                print("Agent terminated from ",str(apartment_obj))
                # ApartmentOp.update_agent(apartment_obj,None)
                if apartment_obj.agent_id == agent_obj.username:
                    ApartmentOp.update_agent(apartment_obj,None)

                    agent_co = agent_obj.company
                    ApartmentOp.update_company(apartment_obj,None)

                    company_users = agent_co.users
                    for i in company_users:
                        print("These are users",company_users)
                        if i.user_group_id == 4:
                            current_user_apartments = fetch_all_apartments_by_user(i)
                            if apartment_obj in current_user_apartments:
                                ApartmentOp.terminate(apartment_obj,i)
                                print("user removed from ",str(apartment_obj))
                            else:
                                print("User did not have access to", str(apartment_obj))


                # # UserOp.update_status(agent_obj,True)
                # if apartment_obj.agency_managed:
                #     agent_obj_co = agent_obj.company_id
                #     ApartmentOp.update_company(apartment_obj,agent_obj_co)

        msg = "Operation completed"
        flash(msg,"success")
        return make_response(jsonify({
                'msg': 'Success',
              'apartmentlist':stringify_list_items(apartments)
            }))
        

class FetchExcel(Resource):
    """read/write Excel xlsx/xlsm files"""
    def get(self):
        return Response(render_template('fetch_excel.html'))

    def post(self):
        excel_file = request.files['file']
        wb = load_workbook(excel_file)
        ws = wb[wb.sheetnames[0]]
        df = pd.read_excel(excel_file)
        # image_loader = SheetImageLoader(ws)
        # image = image_loader.get('A1')
        return Response(render_template('print_excel.html',tables=[df.to_html(classes='data', header="true")]))

class FetchPdf(Resource):
    def post(self):
        pdf_file = request.files['file']
        pass



